import pandas as pd
from datetime import datetime
import unicodedata
import re
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Definir os grupos de veículos
grupos_veiculos = {
    'oeste': ('MATERIAL RODANTE - VLT', ['VLT01', 'VLT02', 'VLT03', 'VLT04', 'VLT05', 'VLT06']),
    'nordeste': ('MATERIAL RODANTE - VLT', ['VLT07', 'VLT08', 'VLT09', 'VLT10', 'VLT11', 'VLT12', 'VLT13']),
    'sobral': ('MATERIAL RODANTE - VLT', ['VLTS02', 'VLTS03', 'VLTS04', 'VLTS05', 'VLTS06']),
    'cariri': ('MATERIAL RODANTE - VLT', ['TRAM1', 'TRAM2', 'VLTC03'])
}

# Base padrão (caminho real onde estão os arquivos de entrada/saída)
base = Path.home() / 'Downloads' / 'Documentos - Farol'

# Tentar diferentes codificações para ler o arquivo. Procurar primeiro em `base`, depois no diretório atual.
encodings = ['latin-1', 'iso-8859-1', 'cp1252', 'utf-8']
csv_candidates = [base / 'SspCompleta.csv', Path('SspCompleta.csv')]

df = None
for candidate in csv_candidates:
    if not candidate.exists():
        continue
    for encoding in encodings:
        try:
            df = pd.read_csv(str(candidate), sep=';', encoding=encoding, low_memory=False)
            break
        except UnicodeDecodeError:
            continue
        except Exception:
            continue
    if df is not None:
        break

if df is None:
    raise Exception("Não foi possível ler o arquivo SspCompleta.csv em nenhum dos locais esperados")

# Verificar valores únicos na coluna de serviço para identificar o nome correto
servico_col = None
for col in df.columns:
    if 'servi' in col.lower() or 'servi' in col.lower():
        servico_col = col
        break

if not servico_col:
    # Procurar por colunas que possam conter os serviços
    for col in df.columns:
        unique_vals = df[col].dropna().unique()
        for val in unique_vals:
            if isinstance(val, str) and ('MANUTEN' in val.upper() or 'PREVENTIVA' in val.upper()):
                servico_col = col
                break
        if servico_col:
            break

# Encontrar a coluna de veículo
veiculo_col = None
for col in df.columns:
    if 'veiculo' in col.lower() or 'veículo' in col.lower():
        veiculo_col = col
        break

if not veiculo_col:
    # Tentar encontrar por padrão
    for col in df.columns:
        if any(term in col.upper() for term in ['VLT', 'TRAM', 'VEIC']):
            veiculo_col = col
            break

if not servico_col or not veiculo_col:
    raise Exception("Não foi possível identificar as colunas necessárias")

# Criar lista de todos os veículos de interesse
todos_veiculos = []
for grupo in grupos_veiculos.values():
    todos_veiculos.extend(grupo[1])

# Filtrar os dados
filtro_veiculos = df[veiculo_col].isin(todos_veiculos)

# Vamos normalizar os nomes de serviço e mapear apenas os que são
# efetivamente "manutenção preventiva diária" ou "manutenção preventiva semanal".
def normalize_text(s):
    s = str(s).upper()
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r'\s+', ' ', s).strip()
    return s

# Mapear valores originais para rótulos canônicos (somente DIÁRIA/SEMANAL)
canonical_map = {}
for serv in df[servico_col].dropna().unique():
    n = normalize_text(serv)
    # requer que seja manutenção e preventiva
    if 'MANUTEN' in n and 'PREVENT' in n:
        if 'DIARIA' in n:
            canonical_map[serv] = 'MANUTENÇÃO PREVENTIVA DIÁRIA'
        elif 'SEMANAL' in n:
            canonical_map[serv] = 'MANUTENÇÃO PREVENTIVA SEMANAL'

servicos_encontrados = list(canonical_map.keys())

if not servicos_encontrados:
    for serv in df[servico_col].dropna().unique():
        n = normalize_text(serv)
        if ('DIARIA' in n or 'SEMANAL' in n) and ('PREVENT' in n or 'MANUTEN' in n):
            canonical = 'MANUTENÇÃO PREVENTIVA DIÁRIA' if 'DIARIA' in n else 'MANUTENÇÃO PREVENTIVA SEMANAL'
            canonical_map[serv] = canonical
            servicos_encontrados.append(serv)

# Criar coluna canônica e aplicar filtro apenas nessas manutenções
df['__servico_canonico'] = df[servico_col].map(canonical_map)
filtro_servicos = df['__servico_canonico'].notna()

# Aplicar filtros
df_filtrado = df[filtro_veiculos & filtro_servicos].copy()

# Substituir a coluna original de serviço pelos valores normalizados (canônicos)
df_filtrado[servico_col] = df_filtrado['__servico_canonico']
if '__servico_canonico' in df_filtrado.columns:
    df_filtrado = df_filtrado.drop(columns=['__servico_canonico'])

if len(df_filtrado) == 0:
    # nenhum registro encontrado — sair silenciosamente
    raise SystemExit(0)
else:
    # Contar manutenções por veículo e tipo de serviço (usar a coluna do serviço já normalizada)
    resultado = df_filtrado.groupby([veiculo_col, servico_col]).size().reset_index(name='Quantidade')

    # Pivotar a tabela para ter serviços como colunas
    resultado_pivot = resultado.pivot_table(
        index=veiculo_col,
        columns=servico_col,
        values='Quantidade',
        fill_value=0
    ).reset_index()
    
    # Renomear colunas para facilitar
    resultado_pivot.columns.name = None
    
    # Garantir que temos as colunas desejadas
    colunas_desejadas = ['MANUTENÇÃO PREVENTIVA DIÁRIA', 'MANUTENÇÃO PREVENTIVA SEMANAL']
    for coluna in colunas_desejadas:
        if coluna not in resultado_pivot.columns:
            resultado_pivot[coluna] = 0
    
    # Reordenar colunas
    colunas_finais = [veiculo_col] + colunas_desejadas
    for col in resultado_pivot.columns:
        if col not in colunas_finais and any(term in str(col).upper() for term in ['MANUTEN', 'PREVENTIVA']):
            colunas_finais.append(col)
    
    resultado_final = resultado_pivot[colunas_finais]
    
    # Adicionar linha de totais
    total_row = resultado_final.sum(numeric_only=True)
    total_row[veiculo_col] = 'TOTAL'
    resultado_final = pd.concat([resultado_final, pd.DataFrame([total_row])], ignore_index=True)
    
    # Salvar em `contagem_safs.xlsx` — adicionar duas abas e aplicar formatação similar a `mining_combined.py`
    # Caminho fixo para o arquivo de saída conforme solicitado
    contagem_path = base / 'dados_resumo.xlsx'

    def write_df_to_sheet(wb, df_out, sheet_name):
        if sheet_name in wb.sheetnames:
            # remove existing sheet to replace
            std = wb[sheet_name]
            wb.remove(std)
        ws = wb.create_sheet(sheet_name)

        # write dataframe rows
        rows = dataframe_to_rows(df_out, index=False, header=True)
        header = next(rows)

        # Header formatting
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_align = Alignment(horizontal='center')

        for col_idx, col_name in enumerate(header, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # write remaining rows
        for r_idx, row in enumerate(rows, 2):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Aplicar larguras específicas pedidas pelo usuário
        try:
            if sheet_name == 'Resumo Manutenções SSP':
                ws.column_dimensions['B'].width = 34
                ws.column_dimensions['C'].width = 34
            elif sheet_name == 'Detalhes Manutenções SSP':
                ws.column_dimensions['B'].width = 34
        except Exception:
            pass

        # Habilitar AutoFilter somente para a aba de detalhes (segunda aba)
        try:
            if sheet_name == 'Detalhes Manutenções SSP':
                last_col = get_column_letter(ws.max_column)
                last_row = ws.max_row
                ws.auto_filter.ref = f"A1:{last_col}{last_row}"
        except Exception:
            pass

        ws.freeze_panes = 'A2'


    # preparar dataframes para escrita
    resumo_df = resultado_final.copy()

    # Detalhes: manter veículo + serviço normalizado e preservar colunas úteis (data da ordem e status), se existirem
    colunas_detalhes = [veiculo_col, servico_col]

    # detectar colunas de data (priorizar nomes comuns)
    date_candidates = [c for c in df_filtrado.columns if 'data' in c.lower()]
    preferred_dates = ['Data Programada', 'Data de Abertura', 'Data de Abertura Saf', 'Data']
    date_col = None
    for p in preferred_dates:
        if p in df_filtrado.columns:
            date_col = p
            break
    if date_col is None and date_candidates:
        date_col = date_candidates[0]

    # detectar coluna de status
    status_col = None
    for c in df_filtrado.columns:
        lc = c.lower()
        if lc == 'status' or 'status' in lc or lc == 'nome status':
            status_col = c
            break

    if date_col and date_col not in colunas_detalhes:
        colunas_detalhes.append(date_col)
    if status_col and status_col not in colunas_detalhes:
        colunas_detalhes.append(status_col)

    detalhes_df = df_filtrado[colunas_detalhes].copy()

    # formatar data para dd/mm/YYYY quando possível
    if date_col and date_col in detalhes_df.columns:
        try:
            detalhes_df[date_col] = pd.to_datetime(detalhes_df[date_col], dayfirst=True, errors='coerce').dt.strftime('%d/%m/%Y')
        except Exception:
            pass

    # load or create workbook no caminho fixo
    if contagem_path.exists():
        wb = load_workbook(contagem_path)
    else:
        # garantir que a pasta exista
        try:
            contagem_path.parent.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass
        wb = Workbook()

    write_df_to_sheet(wb, resumo_df, 'Resumo Manutenções SSP')
    write_df_to_sheet(wb, detalhes_df, 'Detalhes Manutenções SSP')

    # save workbook
    wb.save(contagem_path)

    # Mensagem de confirmação simples
    print(f"Arquivo salvo/atualizado: {contagem_path}")