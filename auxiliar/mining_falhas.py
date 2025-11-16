import pandas as pd
import pathlib
from typing import Dict, List
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

class SAFComFalhasAnalyzer:
    """Classe para análise de SAFs com falhas do equipamento"""
    
    def __init__(self, file_path: str):
        """Inicializa o analisador com o caminho do arquivo CSV"""
        self.downloads_path = pathlib.Path(file_path)
        self.df = self._load_data()
        self.linhas_config = self._get_linhas_config()
        
    def _load_data(self) -> pd.DataFrame:
        """Carrega e filtra os dados do arquivo CSV"""
        try:
            df = pd.read_csv(
                self.downloads_path, 
                encoding='latin-1', 
                sep=';',
                dtype=str
            )
            
            # Converter a coluna de data para datetime
            df["Data de Abertura Saf"] = pd.to_datetime(
                df["Data de Abertura Saf"], 
                dayfirst=True, 
                errors='coerce'
            )
            
            # Remover linhas com datas inválidas
            df = df.dropna(subset=["Data de Abertura Saf"])
            
            # Filtrar apenas SAFs encerradas
            df = df[df['Nome Status'] == 'Encerrada']
            
            # Filtrar apenas falhas do equipamento
            df = df[df['Agente Causador'] == 'FALHA DO EQUIPAMENTO']
            
            print(f"Dados carregados e filtrados com sucesso. {len(df)} falhas de equipamento encontradas.")
            return df
            
        except Exception as e:
            raise Exception(f"Erro ao carregar dados: {str(e)}")
    
    @staticmethod
    def _get_linhas_config() -> Dict[str, List[str]]:
        """Retorna a configuração de veículos por linha"""
        return {
            'Linha Sul': ['TUE'],
            'Linha Oeste': ['VLT01', 'VLT02', 'VLT03', 'VLT04', 'VLT05', 'VLT06'],
            'Linha Parangaba - Mucuripe': ['VLT07', 'VLT08', 'VLT09', 'VLT10', 'VLT11', 'VLT12', 'VLT13'],
            'Linha Sobral': ['VLTS02', 'VLTS03', 'VLTS04', 'VLTS05', 'VLTS06'],
            'Linha Cariri': ['TRAM1', 'TRAM2', 'VLTC03']
        }
    
    def identificar_linha(self, veiculo: str, grupo: str) -> str:
        """Identifica a linha com base no veículo e grupo"""
        if grupo == 'MATERIAL RODANTE - TUE':
            return 'Linha Sul'
        
        for linha, veiculos in self.linhas_config.items():
            if veiculo in veiculos:
                return linha
        
        return 'Outra Linha'
    
    def processar_falhas(self) -> pd.DataFrame:
        """Processa e organiza os dados de falhas do equipamento"""
        print("Processando falhas do equipamento...")
        
        # Criar DataFrame com as colunas necessárias
        falhas_detalhadas = []
        
        for _, row in self.df.iterrows():
            # Identificar a linha
            linha = self.identificar_linha(row['Veiculo'], row['Grupo'])
            
            falha_info = {
                'Data': row['Data de Abertura Saf'].strftime('%d/%m/%Y'),
                'Veículo': row['Veiculo'],
                'Linha': linha,
                'Nível': row['Nível'],
                'Sistema': row['Sistema'],
                'Sub Sistema': row['Sub Sistema'],
                'Agente Causador': row['Agente Causador']
            }
            
            falhas_detalhadas.append(falha_info)
        
        df_falhas = pd.DataFrame(falhas_detalhadas)
        
        # Ordenar por data e linha
        df_falhas['Data'] = pd.to_datetime(df_falhas['Data'], dayfirst=True)
        df_falhas = df_falhas.sort_values(['Data', 'Linha', 'Nível'])
        df_falhas['Data'] = df_falhas['Data'].dt.strftime('%d/%m/%Y')
        
        print(f"Processamento concluído. {len(df_falhas)} registros de falha organizados.")
        return df_falhas
    
    def salvar_na_planilha_existente(self, df_falhas: pd.DataFrame, template_path: str, output_path: str) -> None:
        """Salva os dados detalhados na planilha existente"""
        try:
            # Carregar a planilha template
            wb = load_workbook(template_path)
            
            # Remover aba existente se houver e criar nova
            if 'falhas_equipamento' in wb.sheetnames:
                wb.remove(wb['falhas_equipamento'])
            
            # Criar aba
            ws_detalhes = wb.create_sheet('falhas_equipamento', 0)
            
            # Formatar aba de detalhes
            self._formatar_aba_detalhes(ws_detalhes, df_falhas)
            
            # Salvar a planilha resultante
            wb.save(output_path)
            print(f"Planilha salva com sucesso: {output_path}")
            
        except Exception as e:
            raise Exception(f"Erro ao salvar planilha: {str(e)}")
    
    def _formatar_aba_detalhes(self, worksheet, df_falhas: pd.DataFrame):
        """Formata a aba com os detalhes das falhas"""
        # Cabeçalhos
        headers = ['Data', 'Veículo', 'Linha', 'Nível', 'Sistema', 'Sub Sistema', 'Agente Causador']
        
        # Adicionar cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Adicionar dados
        for row_idx, row_data in enumerate(dataframe_to_rows(df_falhas, index=False, header=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Ajustar largura das colunas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Congelar painel (cabeçalhos)
        worksheet.freeze_panes = 'A2'
        
        print("Aba de detalhes formatada com sucesso.")

def main():
    """Função principal"""
    try:
        # Configurar caminhos dos arquivos
        downloads_path = pathlib.Path.home() / "Downloads" / "Documentos - Farol" / "OsmCompleta.csv"
        template_path = pathlib.Path.home() / "Downloads" / "Documentos - Farol" / "contagem_safs.xlsx"
        output_path = pathlib.Path.home() / "Downloads" / "Documentos - Farol" / "contagem_safs.xlsx"
        
        # Criar diretório de saída se não existir
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Inicializar e executar análise
        analyzer = SAFComFalhasAnalyzer(str(downloads_path))
        df_falhas = analyzer.processar_falhas()
        
        # Salvar resultados na planilha template
        analyzer.salvar_na_planilha_existente(df_falhas, str(template_path), str(output_path))
        
        print("\n" + "="*60)
        print("RELATÓRIO DE FALHAS DO EQUIPAMENTO")
        print("="*60)
        print(f"Total de falhas processadas: {len(df_falhas)}")
        print(f"Arquivo gerado: {output_path}")
        print("="*60)
        
    except Exception as e:
        print(f"Erro durante a execução: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0

if __name__ == "__main__":
    exit(main())