import pandas as pd
import numpy as np
import datetime
import pathlib
from typing import Dict, Tuple
from openpyxl import load_workbook
from openpyxl.styles import Font

class SAFSemFalhasAnalyzer:
    """Classe para análise de SAFs sem falhas baseada no Agente Causador"""
    
    def __init__(self, file_path: str):
        """Inicializa o analisador com o caminho do arquivo CSV"""
        self.downloads_path = pathlib.Path(file_path)
        self.df = self._load_data()
        self.linhas_config = self._get_linhas_config()
        self.contadores = self._inicializar_contadores()
        
    def _load_data(self) -> pd.DataFrame:
        """Carrega os dados do arquivo CSV"""
        try:
            df = pd.read_csv(
                self.downloads_path, 
                encoding='latin-1', 
                sep=';',
                dtype=str
            )
            
            # Converter a coluna de data para datetime
            df["Data de Abertura Saf"] = pd.to_datetime(
                df["Data de Abertura Saf"], 
                dayfirst=True, 
                errors='coerce'
            )
            
            # Remover linhas com datas inválidas
            df = df.dropna(subset=["Data de Abertura Saf"])
            
            # Filtrar apenas SAFs encerradas
            df = df[df['Nome Status'] == 'Encerrada']
            
            # Definir agentes causadores que não caracterizam falha
            agentes_sem_falha = [
                'NÃO CARACTERIZA FALHA',
                'TERCEIROS', 
                'NÃO IDENTIFICADO',
                'TÉRMINO DA VIDA ÚTIL'
            ]
            
            # Filtrar SAFs sem falha
            df = df[df['Agente Causador'].isin(agentes_sem_falha)]
            
            print("Dados carregados e filtrados com sucesso.")
            return df
        except Exception as e:
            raise Exception(f"Erro ao carregar dados: {str(e)}")
    
    @staticmethod
    def _get_linhas_config() -> Dict[str, Tuple[str, list]]:
        """Retorna a configuração de veículos por linha"""
        return {
            'sul': ('MATERIAL RODANTE - TUE', None),
            'oeste': ('MATERIAL RODANTE - VLT', ['VLT01', 'VLT02', 'VLT03', 'VLT04', 'VLT05', 'VLT06']),
            'nordeste': ('MATERIAL RODANTE - VLT', ['VLT07', 'VLT08', 'VLT09', 'VLT10', 'VLT11', 'VLT12', 'VLT13']),
            'sobral': ('MATERIAL RODANTE - VLT', ['VLTS02', 'VLTS03', 'VLTS04', 'VLTS05', 'VLTS06']),
            'cariri': ('MATERIAL RODANTE - VLT', ['TRAM1', 'TRAM2', 'VLTC03'])
        }
    
    def _inicializar_contadores(self) -> Dict[str, pd.DataFrame]:
        """Inicializa os dataframes de contagem com dias como linhas e linhas como colunas"""
        # Dias como índice (1-31), linhas como colunas
        dias = range(1, 32)
        linhas = list(self.linhas_config.keys())
        
        # Inicializar DataFrames vazios com a estrutura desejada
        return {
            'A': pd.DataFrame(0, index=dias, columns=linhas),
            'B': pd.DataFrame(0, index=dias, columns=linhas),
            'C': pd.DataFrame(0, index=dias, columns=linhas)
        }
    
    def filtrar_linha(self, linha: str) -> pd.DataFrame:
        """Filtra o dataframe pela linha especificada"""
        if linha not in self.linhas_config:
            raise ValueError(f"Linha '{linha}' não reconhecida")
        
        grupo, veiculos = self.linhas_config[linha]
        filtro = self.df['Grupo'] == grupo
        
        if veiculos:
            filtro = filtro & self.df['Veiculo'].isin(veiculos)
        
        return self.df[filtro].copy()
    
    def _get_mes_ano_referencia(self, df_linha: pd.DataFrame) -> Tuple[int, int]:
        """Obtém o mês e ano de referência a partir dos dados"""
        if df_linha.empty:
            return datetime.datetime.now().month, datetime.datetime.now().year
            
        # Usar a primeira data válida como referência
        primeira_data = df_linha['Data de Abertura Saf'].iloc[0]
        return primeira_data.month, primeira_data.year
    
    def processar_linha(self, linha: str) -> None:
        """Processa e conta as SAFs sem falha para uma linha específica"""
        print(f"Processando linha {linha.upper()}...")
        
        # Filtrar dados da linha
        df_linha = self.filtrar_linha(linha)
        
        if df_linha.empty:
            print(f"Nenhum dado encontrado para a linha {linha}.")
            return
        
        # Obter mês e ano de referência
        mes, ano = self._get_mes_ano_referencia(df_linha)
        
        # Processar cada dia do mês
        for dia in range(1, 32):
            try:
                # Filtrar por data
                mask_data = (df_linha['Data de Abertura Saf'].dt.day == dia) & \
                           (df_linha['Data de Abertura Saf'].dt.month == mes) & \
                           (df_linha['Data de Abertura Saf'].dt.year == ano)
                
                df_dia = df_linha[mask_data]
                
                if df_dia.empty:
                    continue
                
                # Contar por nível
                for nivel in ['A', 'B', 'C']:
                    count = len(df_dia[df_dia['Nível'] == nivel])
                    # Atualizar o contador (dia como índice, linha como coluna)
                    self.contadores[nivel].at[dia, linha] = count
                
            except ValueError:
                # Dia inválido para o mês (ex: 31 de fevereiro)
                continue
    
    def processar_todas_linhas(self) -> None:
        """Processa todas as linhas definidas na configuração"""
        for linha in self.linhas_config.keys():
            self.processar_linha(linha)

    def salvar_na_planilha_existente(self, template_path: str, output_path: str) -> None:
        """Preenche a planilha template com os dados processados na aba 'sem_falhas'"""
        try:
            # Carregar a planilha template
            wb = load_workbook(template_path)
            ws = wb['sem_falhas']
            
            # Mapeamento das linhas para as colunas na planilha
            mapeamento_linhas = {
                'sul': {'coluna_inicial': 'B', 'nome_planilha': 'TUE'},
                'oeste': {'coluna_inicial': 'F', 'nome_planilha': 'OESTE'},
                'nordeste': {'coluna_inicial': 'J', 'nome_planilha': 'NORDESTE'},
                'sobral': {'coluna_inicial': 'N', 'nome_planilha': 'SOBRAL'},
                'cariri': {'coluna_inicial': 'R', 'nome_planilha': 'CARIRI'}
            }
            
            # Preencher os dados para cada linha
            for linha_interna, info in mapeamento_linhas.items():
                coluna_inicial = info['coluna_inicial']
                
                # Para cada nível (A, B, C) nas colunas consecutivas
                for i, nivel in enumerate(['A', 'B', 'C']):
                    coluna = chr(ord(coluna_inicial) + i)
                    
                    # Preencher os dados para cada dia (linhas 5 a 35 para dias 1-31)
                    for dia in range(1, 32):
                        linha_planilha = dia + 3  # Linha 5 = dia 1, linha 6 = dia 2, etc.
                        valor = self.contadores[nivel].at[dia, linha_interna]
                        ws[f'{coluna}{linha_planilha}'] = valor
            
            # Adicionar cabeçalhos se estiverem vazios
            for linha_interna, info in mapeamento_linhas.items():
                coluna_inicial = info['coluna_inicial']
                nome_planilha = info['nome_planilha']
                
                # Verificar se o cabeçalho está vazio e preencher
                if not ws[f'{coluna_inicial}3'].value:
                    ws[f'{coluna_inicial}3'] = nome_planilha
                
                # Preencher os níveis A, B, C na linha 4
                for i, nivel in enumerate(['A', 'B', 'C']):
                    coluna = chr(ord(coluna_inicial) + i)
                    if not ws[f'{coluna}3'].value:
                        ws[f'{coluna}3'] = nivel
            
            # Adicionar números dos dias na coluna A se estiver vazia
            for dia in range(1, 32):
                linha_planilha = dia + 3
                if not ws[f'A{linha_planilha}'].value:
                    ws[f'A{linha_planilha}'] = dia
            
            # Salvar a planilha resultante
            wb.save(output_path)
            print(f"Planilha preenchida com sucesso: {output_path}")
            
        except Exception as e:
            raise Exception(f"Erro ao preencher planilha: {str(e)}")

def main():
    """Função principal"""
    try:
        # Configurar caminhos dos arquivos
        downloads_path = pathlib.Path.home() / "Downloads" / "Documentos - Farol" / "OsmCompleta.csv"
        template_path = pathlib.Path.home() / "Downloads" / "Documentos - Farol" / "contagem_safs.xlsx"
        output_path = pathlib.Path.home() / "Downloads" / "Documentos - Farol" / "contagem_safs.xlsx"
        
        # Criar diretório de saída se não existir
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Inicializar e executar análise
        analyzer = SAFSemFalhasAnalyzer(str(downloads_path))
        analyzer.processar_todas_linhas()
        
        # Salvar resultados na planilha template
        analyzer.salvar_na_planilha_existente(str(template_path), str(output_path))
        
        print("Processamento concluído com sucesso!")
        print(f"Arquivo gerado: {output_path}")
        
    except Exception as e:
        print(f"Erro durante a execução: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0

if __name__ == "__main__":
    exit(main())