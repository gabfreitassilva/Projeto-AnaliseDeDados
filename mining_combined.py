import pandas as pd
import datetime
import pathlib
from typing import Dict, Tuple, Optional
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


class UnifiedSAFAnalyzer:
    """Analisador unificado para SAFs e SAFs sem falhas.

    Detecta automaticamente colunas relevantes (datas, veículo, grupo, nível)
    e aplica filtros específicos quando `sem_falhas=True`.
    """

    def __init__(self, file_path: str, sem_falhas: bool = False, min_date: Optional[str] = None, max_date: Optional[str] = None):
        self.path = pathlib.Path(file_path)
        self.sem_falhas = sem_falhas
        self._min_date_input = min_date
        self._max_date_input = max_date
        self.df = self._load_data()
        self.linhas_config = self._get_linhas_config()
        self.contadores = self._inicializar_contadores()

    def _load_data(self) -> pd.DataFrame:
        """Carrega CSV e normaliza nomes de colunas essenciais.

        - Tenta reconhecer colunas: data (Data de Abertura / Data de Abertura Saf),
          veículo (Veículo / Veiculo), grupo, nível, Nome Status e Agente Causador.
        - Para `sem_falhas` aplica filtro de status 'Encerrada' e agentes não caracterizam falha.
        """
        try:
            df = pd.read_csv(self.path, encoding="latin-1", sep=';', dtype=str)

            # Normalizar nomes de colunas (remover espaços redundantes)
            df.columns = [col.strip() for col in df.columns]

            # Detectar coluna de data
            if 'Data de Abertura Saf' in df.columns:
                date_col = 'Data de Abertura Saf'
            elif 'Data de Abertura' in df.columns:
                date_col = 'Data de Abertura'
            else:
                # pega a primeira coluna que contenha 'Data' no nome
                found = [c for c in df.columns if 'Data' in c]
                date_col = found[0] if found else None

            if date_col is None:
                raise ValueError('Coluna de data não encontrada no arquivo')

            # Converter data para datetime
            df[date_col] = pd.to_datetime(df[date_col], dayfirst=True, errors='coerce')
            df = df.dropna(subset=[date_col])

            # Aplicar filtro por intervalo de datas se informado
            def _parse_date_input(date_in: Optional[str]) -> Optional[pd.Timestamp]:
                if not date_in:
                    return None
                for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d%m%Y"):
                    try:
                        return pd.to_datetime(datetime.datetime.strptime(date_in, fmt))
                    except Exception:
                        continue
                # fallback para tentativa genérica
                try:
                    return pd.to_datetime(date_in, dayfirst=True, errors='coerce')
                except Exception:
                    return None

            min_dt = _parse_date_input(self._min_date_input)
            max_dt = _parse_date_input(self._max_date_input)

            if min_dt is not None or max_dt is not None:
                before_dt = len(df)
                if min_dt is not None:
                    df = df[df[date_col] >= min_dt]
                if max_dt is not None:
                    df = df[df[date_col] <= max_dt]
                after_dt = len(df)
                print(f"Filtragem por intervalo de datas aplicada: {before_dt} -> {after_dt} registros mantidos.")

            # Filtrar apenas grupos relevantes (somente MATERIAL RODANTE - TUE e MATERIAL RODANTE - VLT)
            grupos_permitidos = ['MATERIAL RODANTE - TUE', 'MATERIAL RODANTE - VLT']
            if 'Grupo' in df.columns:
                before = len(df)
                df = df[df['Grupo'].isin(grupos_permitidos)]
                after = len(df)
                print(f"Filtragem por Grupo aplicada: {before} -> {after} registros mantidos.")
            else:
                print("Aviso: coluna 'Grupo' não encontrada; nenhum filtro por Grupo aplicado.")

            # Garantir nomes de colunas consistentes
            if 'Veículo' in df.columns and 'Veiculo' not in df.columns:
                df = df.rename(columns={'Veículo': 'Veiculo'})

            # Filtrar SAFs encerradas e sem-falhas quando solicitado
            if self.sem_falhas:
                # Nome Status
                if 'Nome Status' in df.columns:
                    df = df[df['Nome Status'] == 'Encerrada']

                agentes_sem_falha = [
                    'NÃO CARACTERIZA FALHA',
                    'TERCEIROS',
                    'NÃO IDENTIFICADO',
                    'TÉRMINO DA VIDA ÚTIL',
                    'MANUTENÇÃO',
                    'INEXISTENTE',
                    'ERRO OPERACIONAL',
                    'ERRO NO PROCESSAMENTO',
                    'DEGRADAÇOES ADVINDAS DE CAUSAS EXTERNAS',
                    'ATUAÇÃO DE PROTEÇÃO'
                ]

                if 'Agente Causador' in df.columns:
                    df = df[df['Agente Causador'].isin(agentes_sem_falha)]

            # Armazenar o nome efetivo da coluna de data para uso posterior
            self._date_col = date_col
            print(f"Dados carregados ({len(df)} registros) de: {self.path.name}")
            return df
        except Exception as e:
            raise Exception(f"Erro ao carregar dados: {e}")

    @staticmethod
    def _get_linhas_config() -> Dict[str, Tuple[str, Optional[list]]]:
        return {
            'sul': ('MATERIAL RODANTE - TUE', None),
            'oeste': ('MATERIAL RODANTE - VLT', ['VLT01', 'VLT02', 'VLT03', 'VLT04', 'VLT05', 'VLT06']),
            'nordeste': ('MATERIAL RODANTE - VLT', ['VLT07', 'VLT08', 'VLT09', 'VLT10', 'VLT11', 'VLT12', 'VLT13']),
            'sobral': ('MATERIAL RODANTE - VLT', ['VLTS02', 'VLTS03', 'VLTS04', 'VLTS05', 'VLTS06']),
            'cariri': ('MATERIAL RODANTE - VLT', ['TRAM1', 'TRAM2', 'VLTC03'])
        }

    def _inicializar_contadores(self) -> Dict[str, pd.DataFrame]:
        dias = range(1, 32)
        linhas = list(self.linhas_config.keys())
        return {
            'A': pd.DataFrame(0, index=dias, columns=linhas),
            'B': pd.DataFrame(0, index=dias, columns=linhas),
            'C': pd.DataFrame(0, index=dias, columns=linhas)
        }

    def filtrar_linha(self, linha: str) -> pd.DataFrame:
        if linha not in self.linhas_config:
            raise ValueError(f"Linha '{linha}' não reconhecida")

        grupo, veiculos = self.linhas_config[linha]
        if 'Grupo' not in self.df.columns:
            raise ValueError("Coluna 'Grupo' não encontrada no dataset")

        filtro = self.df['Grupo'] == grupo

        if veiculos:
            # algumas fontes usam 'Veiculo' ou outra coluna, assumimos 'Veiculo'
            if 'Veiculo' not in self.df.columns:
                raise ValueError("Coluna 'Veiculo' não encontrada no dataset")
            filtro = filtro & self.df['Veiculo'].isin(veiculos)

        return self.df[filtro].copy()

    def _get_mes_ano_referencia(self, df_linha: pd.DataFrame) -> Tuple[int, int]:
        if df_linha.empty:
            now = datetime.datetime.now()
            return now.month, now.year
        primeira_data = df_linha[self._date_col].iloc[0]
        return primeira_data.month, primeira_data.year

    def processar_linha(self, linha: str) -> None:
        print(f"Processando linha {linha.upper()} ({'sem_falhas' if self.sem_falhas else 'safs'})...")
        df_linha = self.filtrar_linha(linha)

        if df_linha.empty:
            print(f"Nenhum dado encontrado para a linha {linha}.")
            return

        mes, ano = self._get_mes_ano_referencia(df_linha)

        for dia in range(1, 32):
            try:
                mask_data = (
                    (df_linha[self._date_col].dt.day == dia) &
                    (df_linha[self._date_col].dt.month == mes) &
                    (df_linha[self._date_col].dt.year == ano)
                )

                df_dia = df_linha[mask_data]
                if df_dia.empty:
                    continue

                for nivel in ['A', 'B', 'C']:
                    count = len(df_dia[df_dia.get('Nível') == nivel]) if 'Nível' in df_dia.columns else 0
                    self.contadores[nivel].at[dia, linha] = count

            except ValueError:
                continue

    def processar_todas_linhas(self) -> None:
        for linha in self.linhas_config.keys():
            self.processar_linha(linha)

    def salvar_na_planilha_existente(self, template_path: str, output_path: str, sheet_name: Optional[str] = None) -> None:
        try:
            wb = load_workbook(template_path)
            if sheet_name is None:
                sheet_name = 'sem_falhas' if self.sem_falhas else 'safs_diarias'

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Planilha '{sheet_name}' não encontrada no template")

            ws = wb[sheet_name]

            mapeamento_linhas = {
                'sul': {'coluna_inicial': 'B', 'nome_planilha': 'TUE'},
                'oeste': {'coluna_inicial': 'F', 'nome_planilha': 'OESTE'},
                'nordeste': {'coluna_inicial': 'J', 'nome_planilha': 'NORDESTE'},
                'sobral': {'coluna_inicial': 'N', 'nome_planilha': 'SOBRAL'},
                'cariri': {'coluna_inicial': 'R', 'nome_planilha': 'CARIRI'}
            }

            # Preencher contadores
            for linha_interna, info in mapeamento_linhas.items():
                coluna_inicial = info['coluna_inicial']
                for i, nivel in enumerate(['A', 'B', 'C']):
                    coluna = chr(ord(coluna_inicial) + i)
                    for dia in range(1, 32):
                        linha_planilha = dia + 3
                        valor = int(self.contadores[nivel].at[dia, linha_interna])
                        ws[f'{coluna}{linha_planilha}'] = valor

            # Cabeçalhos e dias
            for linha_interna, info in mapeamento_linhas.items():
                coluna_inicial = info['coluna_inicial']
                nome_planilha = info['nome_planilha']
                if not ws[f'{coluna_inicial}3'].value:
                    ws[f'{coluna_inicial}3'] = nome_planilha
                for i, nivel in enumerate(['A', 'B', 'C']):
                    coluna = chr(ord(coluna_inicial) + i)
                    if not ws[f'{coluna}3'].value:
                        ws[f'{coluna}3'] = nivel

            for dia in range(1, 32):
                linha_planilha = dia + 3
                if not ws[f'A{linha_planilha}'].value:
                    ws[f'A{linha_planilha}'] = dia

            wb.save(output_path)
            print(f"Planilha preenchida com sucesso: {output_path} (aba: {sheet_name})")

        except Exception as e:
            raise Exception(f"Erro ao preencher planilha: {e}")


class SAFComFalhasAnalyzer:
    """Classe para análise de SAFs com falhas do equipamento"""

    def __init__(self, file_path: str, min_date: Optional[str] = None, max_date: Optional[str] = None):
        self.downloads_path = pathlib.Path(file_path)
        self._min_date_input = min_date
        self._max_date_input = max_date
        self.df = self._load_data()
        self.linhas_config = self._get_linhas_config()

    def _load_data(self) -> pd.DataFrame:
        try:
            df = pd.read_csv(
                self.downloads_path,
                encoding='latin-1',
                sep=';',
                dtype=str
            )

            df["Data de Abertura Saf"] = pd.to_datetime(
                df["Data de Abertura Saf"],
                dayfirst=True,
                errors='coerce'
            )

            df = df.dropna(subset=["Data de Abertura Saf"])

            # Aplicar filtro por intervalo de datas se informado
            def _parse_date_input(date_in: Optional[str]) -> Optional[pd.Timestamp]:
                if not date_in:
                    return None
                for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d%m%Y"):
                    try:
                        return pd.to_datetime(datetime.datetime.strptime(date_in, fmt))
                    except Exception:
                        continue
                try:
                    return pd.to_datetime(date_in, dayfirst=True, errors='coerce')
                except Exception:
                    return None

            min_dt = _parse_date_input(self._min_date_input)
            max_dt = _parse_date_input(self._max_date_input)

            if min_dt is not None or max_dt is not None:
                before_dt = len(df)
                if min_dt is not None:
                    df = df[df['Data de Abertura Saf'] >= min_dt]
                if max_dt is not None:
                    df = df[df['Data de Abertura Saf'] <= max_dt]
                after_dt = len(df)
                print(f"Filtragem por intervalo de datas aplicada (falhas): {before_dt} -> {after_dt} registros mantidos.")

            # Filtrar apenas grupos relevantes (somente MATERIAL RODANTE - TUE e MATERIAL RODANTE - VLT)
            grupos_permitidos = ['MATERIAL RODANTE - TUE', 'MATERIAL RODANTE - VLT']
            if 'Grupo' in df.columns:
                before = len(df)
                df = df[df['Grupo'].isin(grupos_permitidos)]
                after = len(df)
                print(f"Filtragem por Grupo aplicada (falhas): {before} -> {after} registros mantidos.")
            else:
                print("Aviso: coluna 'Grupo' não encontrada no arquivo de falhas; nenhum filtro por Grupo aplicado.")

            df = df[df['Nome Status'] == 'Encerrada']
            df = df[df['Agente Causador'] == 'FALHA DO EQUIPAMENTO']

            print(f"Dados carregados e filtrados com sucesso. {len(df)} falhas de equipamento encontradas.")
            return df

        except Exception as e:
            raise Exception(f"Erro ao carregar dados: {str(e)}")

    @staticmethod
    def _get_linhas_config() -> Dict[str, list]:
        return {
            'Linha Sul': ['TUE'],
            'Linha Oeste': ['VLT01', 'VLT02', 'VLT03', 'VLT04', 'VLT05', 'VLT06'],
            'Linha Parangaba - Mucuripe': ['VLT07', 'VLT08', 'VLT09', 'VLT10', 'VLT11', 'VLT12', 'VLT13'],
            'Linha Sobral': ['VLTS02', 'VLTS03', 'VLTS04', 'VLTS05', 'VLTS06'],
            'Linha Cariri': ['TRAM1', 'TRAM2', 'VLTC03']
        }

    def identificar_linha(self, veiculo: str, grupo: str) -> str:
        if grupo == 'MATERIAL RODANTE - TUE':
            return 'Linha Sul'

        for linha, veiculos in self.linhas_config.items():
            if veiculo in veiculos:
                return linha

        return 'Outra Linha'

    def processar_falhas(self) -> pd.DataFrame:
        print("Processando falhas do equipamento...")

        falhas_detalhadas = []

        for _, row in self.df.iterrows():
            linha = self.identificar_linha(row['Veiculo'], row['Grupo'])

            falha_info = {
                'Data': row['Data de Abertura Saf'].strftime('%d/%m/%Y'),
                'Veículo': row['Veiculo'],
                'Linha': linha,
                'Nível': row['Nível'],
                'Sistema': row.get('Sistema'),
                'Sub Sistema': row.get('Sub Sistema'),
                'Agente Causador': row.get('Agente Causador')
            }

            falhas_detalhadas.append(falha_info)

        df_falhas = pd.DataFrame(falhas_detalhadas)
        df_falhas['Data'] = pd.to_datetime(df_falhas['Data'], dayfirst=True)
        df_falhas = df_falhas.sort_values(['Data', 'Linha', 'Nível'])
        df_falhas['Data'] = df_falhas['Data'].dt.strftime('%d/%m/%Y')

        print(f"Processamento concluído. {len(df_falhas)} registros de falha organizados.")
        return df_falhas

    def salvar_na_planilha_existente(self, df_falhas: pd.DataFrame, template_path: str, output_path: str) -> None:
        try:
            wb = load_workbook(template_path)

            if 'falhas_equipamento' in wb.sheetnames:
                wb.remove(wb['falhas_equipamento'])

            ws_detalhes = wb.create_sheet('falhas_equipamento', 0)

            self._formatar_aba_detalhes(ws_detalhes, df_falhas)

            wb.save(output_path)
            print(f"Planilha salva com sucesso: {output_path}")

        except Exception as e:
            raise Exception(f"Erro ao salvar planilha: {str(e)}")

    def _formatar_aba_detalhes(self, worksheet, df_falhas: pd.DataFrame):
        headers = ['Data', 'Veículo', 'Linha', 'Nível', 'Sistema', 'Sub Sistema', 'Agente Causador']

        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row_data in enumerate(dataframe_to_rows(df_falhas, index=False, header=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        worksheet.freeze_panes = 'A2'
        print("Aba de detalhes formatada com sucesso.")


def main():
    try:
        base = pathlib.Path.home() / "Downloads" / "Documentos - Farol"
        template_path = base / "contagem_safs.xlsx"
        output_path = base / "contagem_safs.xlsx"

        # Intervalo de datas (mude conforme necessário; use formatos: '01-10-2025', '01102025' ou '01/10/2025')
        min_date_str = '01-10-2025'
        max_date_str = '31-10-2025'

        # Arquivos esperados
        saf_file = base / "SafCompleta.csv"
        osm_file = base / "OsmCompleta.csv"

        processed_any = False

        if saf_file.exists():
            analyzer = UnifiedSAFAnalyzer(str(saf_file), sem_falhas=False, min_date=min_date_str, max_date=max_date_str)
            analyzer.processar_todas_linhas()
            analyzer.salvar_na_planilha_existente(str(template_path), str(output_path), sheet_name='safs_diarias')
            processed_any = True

        if osm_file.exists():
            analyzer2 = UnifiedSAFAnalyzer(str(osm_file), sem_falhas=True, min_date=min_date_str, max_date=max_date_str)
            analyzer2.processar_todas_linhas()
            analyzer2.salvar_na_planilha_existente(str(template_path), str(output_path), sheet_name='sem_falhas')
            processed_any = True

        # Processar falhas do equipamento (detalhado) se existir o arquivo OSM
        if osm_file.exists():
            falhas_analyzer = SAFComFalhasAnalyzer(str(osm_file), min_date=min_date_str, max_date=max_date_str)
            df_falhas = falhas_analyzer.processar_falhas()
            falhas_analyzer.salvar_na_planilha_existente(df_falhas, str(template_path), str(output_path))
            processed_any = True

        if not processed_any:
            print("Nenhum arquivo 'SafCompleta.csv' ou 'OsmCompleta.csv' encontrado no diretório esperado:", base)
            return 1

        print("Processamento concluído com sucesso!")
        return 0

    except Exception as e:
        print(f"Erro durante a execução: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    raise SystemExit(main())
