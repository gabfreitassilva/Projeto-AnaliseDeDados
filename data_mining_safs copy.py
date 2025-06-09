import pandas as pd
import numpy as np
import datetime
import chardet
from openpyxl import load_workbook

# Detecta a codificação do arquivo
with open('sheets/SafCompleta.csv', 'rb') as f:
    rawdata = f.read(10000)  # Lê amostra inicial
    result = chardet.detect(rawdata)

# Efetua a leitura do arquivo, armazenando em um dataframe
df = pd.read_csv('sheets/SafCompleta.csv', encoding=result['encoding'], sep=';')

# Filtrando as SAFs por linha
def linha_trem(dataframe, linha): # Função para filtrar e retornar a linha de VLT desejada
    filtro_material_rodante = dataframe['Grupo'] == 'MATERIAL RODANTE - VLT'
    if not filtro_material_rodante.empty:    
        match linha:
            case 'sul':
                linha_sul = dataframe[dataframe['Grupo'] == 'MATERIAL RODANTE - TUE']
                return linha_sul
            case 'oeste':
                linha_oeste = dataframe[filtro_material_rodante & dataframe['Veículo'].isin(['VLT01', 'VLT02', 'VLT03', 'VLT04', 'VLT05', 'VLT06'])]
                return linha_oeste
            case 'nordeste':
                linha_nordeste = dataframe[filtro_material_rodante & dataframe['Veículo'].isin(['VLT07', 'VLT08', 'VLT09', 'VLT10', 'VLT11', 'VLT12', 'VLT13'])]
                return linha_nordeste
            case 'sobral':
                linha_sobral = dataframe[filtro_material_rodante & dataframe['Veículo'].isin(['VLTS02', 'VLTS03', 'VLTS04', 'VLTS05', 'VLTS06'])]
                return linha_sobral
            case 'cariri':
                linha_cariri = dataframe[filtro_material_rodante & dataframe['Veículo'].isin(['TRAM1', 'TRAM2', 'VLTC03'])]
                return linha_cariri
    else:
        print("Dataframe sem dados...")

countA = np.zeros(32, dtype=int)
countB = np.zeros(32, dtype=int)
countC = np.zeros(32, dtype=int)

# Filtrando e fazendo a contagem de SAFs por dia [TUE]
def contagem_safs(dataframe, linha):
    match linha:
        case 'sul':
            dataframe = linha_trem(dataframe, 'sul').copy()
            if not dataframe.empty:
                dataframe["Data de Abertura"] = pd.to_datetime(dataframe["Data de Abertura"], dayfirst=True, format="mixed")
                dataframe["Apenas_Data"] = dataframe["Data de Abertura"].dt.strftime('%d/%m/%Y')

                # Verificar o formato real das datas na coluna
                primeira_data = dataframe['Apenas_Data'].iloc[0]
                print(f"Formato da data de exemplo: {primeira_data}")

                # Extrair mês e ano de uma data válida
                try:
                    partes = primeira_data.split('/')
                    dia = int(partes[0])
                    mes = int(partes[1])
                    ano = int(partes[2])
                except (IndexError, ValueError):
                    mes = datetime.now().month  # Valor padrão se falhar
                    ano = datetime.now().year   # Valor padrão se falhar

                # Iterar pelos dias do mês (de 1 a 31)
                for dia in range(1, 32):
                    data_str = f"{dia:02d}/{mes:02d}/{ano}"
                    filtro_data = dataframe['Apenas_Data'] == data_str
                    df_filtrado = dataframe[filtro_data]
                # Contar os níveis
                    countA[dia] = len(df_filtrado[df_filtrado['Nível'] == 'A'])
                    countB[dia] = len(df_filtrado[df_filtrado['Nível'] == 'B'])
                    countC[dia] = len(df_filtrado[df_filtrado['Nível'] == 'C'])
                    
                # Verifica se o DataFrame filtrado não está vazio
                    if not df_filtrado.empty: 
                        print(f" Dados para {data_str}: ".center(70, '='))
                        print(df_filtrado[['Nível', 'Grupo', 'Veículo', 'Apenas_Data', 'Status']])
                        print(f"Dia {dia}: A = {countA[dia]}, B = {countB[dia]}, C = {countC[dia]}.")
                        print(70*'=', '\n')

linha_sul = linha_trem(df, 'sul')
# contagem_safs(linha_sul, 'sul')

# Carrega o excel que vou editar
wb = load_workbook("sheets/Farol Diário_Acompanhamento de Atividades 14_05_2025.xlsx", read_only=False, rich_text=True)
sheet = wb['TUE'] # Seleciona a aba da planilha que quero trabalhar

# Teste de leitura e escrita de E54 até o AI54
n=5
for coluna in sheet.iter_cols(min_row=55, max_row=55, min_col=5, max_col=35, values_only=True):
    print(coluna[0])
    sheet.cell(row=55, column=n).value = '-'
    n+=1

wb.save("sheets/Farol_copy.xlsx") # ao final, salvo a planilha modificada como uma cópia
